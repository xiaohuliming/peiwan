"""
数据导出服务 (基于 openpyxl)
"""
import io
from decimal import Decimal
from datetime import datetime, date, time
from sqlalchemy.orm import joinedload

from app.extensions import db
from app.utils.time_utils import fmt_dt

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from app.models.user import User
from app.models.order import Order
from app.models.gift import GiftOrder
from app.models.finance import WithdrawRequest, BalanceLog, CommissionLog
from app.models.clock import ClockRecord
from app.models.project import ProjectItem
from app.models.intimacy import Intimacy
from app.models.operation_log import OperationLog
from app.models.lottery import Lottery, LotteryWinner


EXPORT_SECTION_OPTIONS = [
    {'key': 'users', 'label': '用户详情', 'sheet': '用户详情', 'desc': '用户基础资料、钱包、身份标签'},
    {'key': 'orders', 'label': '订单数据', 'sheet': '订单数据', 'desc': '订单全流程明细（老板/陪玩/客服/金额/时间）'},
    {'key': 'gifts', 'label': '礼物订单', 'sheet': '礼物订单', 'desc': '礼物赠送与退款明细'},
    {'key': 'staff', 'label': '客服数据', 'sheet': '客服数据', 'desc': '客服/管理员维度运营数据'},
    {'key': 'players', 'label': '陪玩数据', 'sheet': '陪玩数据', 'desc': '陪玩接单/收益/提现汇总'},
    {'key': 'bosses', 'label': '老板数据', 'sheet': '老板数据', 'desc': '老板充值/消费/下单/赠礼汇总'},
    {'key': 'withdrawals', 'label': '提现数据', 'sheet': '提现数据', 'desc': '提现申请与审核流转'},
    {'key': 'balance_logs', 'label': '嗯呢币流水', 'sheet': '嗯呢币流水', 'desc': '嗯呢币余额变动明细'},
    {'key': 'commission_logs', 'label': '小猪粮流水', 'sheet': '小猪粮流水', 'desc': '小猪粮收益/提现/扣减明细'},
    {'key': 'clocks', 'label': '打卡数据', 'sheet': '打卡数据', 'desc': '客服/管理员打卡明细'},
    {'key': 'intimacies', 'label': '亲密度关系', 'sheet': '亲密度关系', 'desc': '老板-陪玩亲密度关系'},
    {'key': 'lotteries', 'label': '抽奖活动', 'sheet': '抽奖活动', 'desc': '抽奖配置与活动状态'},
    {'key': 'lottery_winners', 'label': '抽奖中奖记录', 'sheet': '抽奖中奖记录', 'desc': '抽奖中奖名单明细'},
    {'key': 'operation_logs', 'label': '操作日志', 'sheet': '操作日志', 'desc': '后台操作审计日志'},
    {'key': 'projects', 'label': '项目配置', 'sheet': '项目配置', 'desc': '游戏项目与档位价格'},
]

EXPORT_SECTION_KEYS = {item['key'] for item in EXPORT_SECTION_OPTIONS}


def _style_header(ws, headers):
    """样式化表头"""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(str(h)) * 2, 12)


def _to_export_str(value):
    """将不同类型值安全转换为可导出的字符串。"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, time):
        return value.strftime('%H:%M:%S')
    if isinstance(value, Decimal):
        return format(value, 'f')
    if isinstance(value, (dict, list, tuple, set)):
        import json
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def export_users(query=None):
    """导出用户列表"""
    if not HAS_OPENPYXL:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = '用户列表'
    headers = ['ID', '用户名', '昵称', '陪玩昵称', 'KOOK ID', '角色', '嗯呢币', '赠金', '小猪粮', '冻结小猪粮',
               '经验值', 'VIP等级', '注册时间']
    _style_header(ws, headers)

    users = query.all() if query else User.query.all()
    for i, u in enumerate(users, 2):
        ws.cell(row=i, column=1, value=u.id)
        ws.cell(row=i, column=2, value=u.username)
        ws.cell(row=i, column=3, value=u.nickname)
        ws.cell(row=i, column=4, value=u.player_nickname)
        ws.cell(row=i, column=5, value=u.kook_id)
        ws.cell(row=i, column=6, value=u.role_name)
        ws.cell(row=i, column=7, value=float(u.m_coin))
        ws.cell(row=i, column=8, value=float(u.m_coin_gift))
        ws.cell(row=i, column=9, value=float(u.m_bean))
        ws.cell(row=i, column=10, value=float(u.m_bean_frozen))
        ws.cell(row=i, column=11, value=u.experience)
        ws.cell(row=i, column=12, value=u.vip_level)
        ws.cell(row=i, column=13, value=u.created_at.strftime('%Y-%m-%d %H:%M') if u.created_at else '')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_orders(query=None):
    """导出订单列表"""
    if not HAS_OPENPYXL:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = '订单列表'
    headers = ['订单号', '老板', '陪玩', '项目', '时长', '总价', '陪玩收益', '平台收入',
               '状态', '冻结', '创建时间', '客服']
    _style_header(ws, headers)

    orders = query.all() if query else Order.query.all()
    for i, o in enumerate(orders, 2):
        ws.cell(row=i, column=1, value=o.order_no)
        ws.cell(row=i, column=2, value=o.boss.nickname if o.boss else '')
        ws.cell(row=i, column=3, value=(o.player.player_nickname or o.player.nickname) if o.player else '')
        ws.cell(row=i, column=4, value=o.project_display)
        ws.cell(row=i, column=5, value=float(o.duration or 0))
        ws.cell(row=i, column=6, value=float(o.total_price or 0))
        ws.cell(row=i, column=7, value=float(o.player_earning or 0))
        ws.cell(row=i, column=8, value=float(o.shop_earning or 0))
        ws.cell(row=i, column=9, value=o.status_label)
        ws.cell(row=i, column=10, value='是' if o.is_frozen else '否')
        ws.cell(row=i, column=11, value=o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else '')
        ws.cell(row=i, column=12, value=o.staff.staff_display_name if o.staff else '')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_gift_orders(query=None):
    """导出礼物记录"""
    if not HAS_OPENPYXL:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = '礼物记录'
    headers = ['ID', '老板', '陪玩', '礼物', '数量', '总价', '陪玩收益', '状态', '时间', '客服']
    _style_header(ws, headers)

    records = query.all() if query else GiftOrder.query.all()
    for i, g in enumerate(records, 2):
        ws.cell(row=i, column=1, value=g.id)
        ws.cell(row=i, column=2, value=g.boss.nickname if g.boss else '')
        ws.cell(row=i, column=3, value=(g.player.player_nickname or g.player.nickname) if g.player else '')
        ws.cell(row=i, column=4, value=g.gift.name if g.gift else '')
        ws.cell(row=i, column=5, value=g.quantity)
        ws.cell(row=i, column=6, value=float(g.total_price))
        ws.cell(row=i, column=7, value=float(g.player_earning))
        ws.cell(row=i, column=8, value=g.status)
        ws.cell(row=i, column=9, value=g.created_at.strftime('%Y-%m-%d %H:%M') if g.created_at else '')
        ws.cell(row=i, column=10, value=g.staff.staff_display_name if g.staff else '')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_withdrawals(query=None):
    """导出提现记录"""
    if not HAS_OPENPYXL:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = '提现记录'
    headers = ['ID', '用户', '金额', '状态', '申请时间', '审批人', '审批时间', '备注']
    _style_header(ws, headers)

    records = query.all() if query else WithdrawRequest.query.all()
    for i, w in enumerate(records, 2):
        ws.cell(row=i, column=1, value=w.id)
        ws.cell(row=i, column=2, value=w.user.nickname if w.user else '')
        ws.cell(row=i, column=3, value=float(w.amount))
        ws.cell(row=i, column=4, value=w.status)
        ws.cell(row=i, column=5, value=w.created_at.strftime('%Y-%m-%d %H:%M') if w.created_at else '')
        ws.cell(row=i, column=6, value=w.auditor.nickname if w.auditor else '')
        ws.cell(row=i, column=7, value=w.audit_at.strftime('%Y-%m-%d %H:%M') if w.audit_at else '')
        ws.cell(row=i, column=8, value=w.audit_remark or '')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_clock_records(query=None):
    """导出打卡记录"""
    if not HAS_OPENPYXL:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = '打卡记录'
    headers = ['ID', '用户', '上班时间', '下班时间', '时长(小时)', '状态']
    _style_header(ws, headers)

    records = query.all() if query else ClockRecord.query.all()
    for i, c in enumerate(records, 2):
        ws.cell(row=i, column=1, value=c.id)
        ws.cell(row=i, column=2, value=c.user.nickname if c.user else '')
        ws.cell(row=i, column=3, value=c.clock_in.strftime('%Y-%m-%d %H:%M') if c.clock_in else '')
        ws.cell(row=i, column=4, value=c.clock_out.strftime('%Y-%m-%d %H:%M') if c.clock_out else '')
        ws.cell(row=i, column=5, value=round((c.duration_minutes or 0) / 60, 2))
        ws.cell(row=i, column=6, value=c.status or '')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _safe_sheet_name(raw_name, used_names):
    """生成合法且唯一的 Excel sheet 名。"""
    if not raw_name:
        raw_name = 'sheet'
    invalid = set('[]:*?/\\')
    base = ''.join('_' if ch in invalid else ch for ch in str(raw_name))
    base = (base or 'sheet')[:31]

    name = base
    idx = 1
    while name in used_names:
        suffix = f'_{idx}'
        name = f'{base[:31-len(suffix)]}{suffix}'
        idx += 1
    used_names.add(name)
    return name


def _to_float(value):
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _main_identity_label(role):
    return {
        'god': '老板',
        'player': '陪玩',
        'staff': '客服',
        'admin': '管理员',
        'superadmin': '高级管理员',
    }.get(role, role or '')


def _identity_summary(user):
    labels = []
    seen = set()

    primary = _main_identity_label(user.role)
    if primary and primary not in seen:
        labels.append(primary)
        seen.add(primary)

    for tag in (user.tag_list or []):
        if tag not in seen:
            labels.append(tag)
            seen.add(tag)

    return ' + '.join(labels)


def _display_name(user, prefer_player=False):
    if not user:
        return ''
    if prefer_player:
        for candidate in (user.player_nickname, user.nickname, user.kook_username, user.username):
            if candidate:
                return candidate
    else:
        for candidate in (user.nickname, user.player_nickname, user.kook_username, user.username):
            if candidate:
                return candidate
    return str(user.id)


def _append_sheet(wb, used_names, title, headers, rows):
    sheet_name = _safe_sheet_name(title, used_names)
    ws = wb.create_sheet(title=sheet_name)
    _style_header(ws, headers)

    max_lens = [len(str(h)) for h in headers]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            text = _to_export_str(val)
            ws.cell(row=r_idx, column=c_idx, value=text)
            if c_idx - 1 < len(max_lens):
                max_lens[c_idx - 1] = max(max_lens[c_idx - 1], len(text))

    for idx, max_len in enumerate(max_lens, start=1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col_letter].width = min(max(int(max_len * 1.4), 12), 60)

    return sheet_name, len(rows)


def export_all_tables_workbook(include_sections=None, date_from=None, date_to=None):
    """按业务关系导出可读型全量数据工作簿。支持 date_from/date_to 日期范围筛选。"""
    if not HAS_OPENPYXL:
        return None

    wb = Workbook()
    info_ws = wb.active
    info_ws.title = '导出说明'
    used_sheet_names = {info_ws.title}
    sheets_summary = []

    # 日期范围辅助 —— 用户输入北京时间，DB 存 UTC，需转换
    from datetime import timezone, timedelta
    _bj_tz = timezone(timedelta(hours=8))
    _dt_from = None
    _dt_to = None
    if date_from:
        try:
            # 北京时间当天 00:00:00 → UTC
            _dt_from = datetime.strptime(date_from, '%Y-%m-%d').replace(tzinfo=_bj_tz).astimezone(timezone.utc).replace(tzinfo=None)
        except ValueError:
            pass
    if date_to:
        try:
            # 北京时间当天 23:59:59 → UTC
            _dt_to = datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S').replace(tzinfo=_bj_tz).astimezone(timezone.utc).replace(tzinfo=None)
        except ValueError:
            pass

    users = User.query.order_by(User.id.asc()).all()
    user_map = {u.id: u for u in users}

    order_q = (
        Order.query
        .options(
            joinedload(Order.boss),
            joinedload(Order.player),
            joinedload(Order.staff),
            joinedload(Order.project_item).joinedload(ProjectItem.project),
        )
    )
    if _dt_from:
        order_q = order_q.filter(Order.created_at >= _dt_from)
    if _dt_to:
        order_q = order_q.filter(Order.created_at <= _dt_to)
    orders = order_q.order_by(Order.created_at.desc()).all()

    gift_q = (
        GiftOrder.query
        .options(
            joinedload(GiftOrder.boss),
            joinedload(GiftOrder.player),
            joinedload(GiftOrder.staff),
            joinedload(GiftOrder.gift),
        )
    )
    if _dt_from:
        gift_q = gift_q.filter(GiftOrder.created_at >= _dt_from)
    if _dt_to:
        gift_q = gift_q.filter(GiftOrder.created_at <= _dt_to)
    gift_orders = gift_q.order_by(GiftOrder.created_at.desc()).all()

    wr_q = (
        WithdrawRequest.query
        .options(
            joinedload(WithdrawRequest.user),
            joinedload(WithdrawRequest.auditor),
        )
    )
    if _dt_from:
        wr_q = wr_q.filter(WithdrawRequest.created_at >= _dt_from)
    if _dt_to:
        wr_q = wr_q.filter(WithdrawRequest.created_at <= _dt_to)
    withdraw_requests = wr_q.order_by(WithdrawRequest.created_at.desc()).all()

    bl_q = BalanceLog.query.options(joinedload(BalanceLog.user))
    if _dt_from:
        bl_q = bl_q.filter(BalanceLog.created_at >= _dt_from)
    if _dt_to:
        bl_q = bl_q.filter(BalanceLog.created_at <= _dt_to)
    balance_logs = bl_q.order_by(BalanceLog.created_at.desc()).all()

    cl_q = (
        CommissionLog.query
        .options(
            joinedload(CommissionLog.user),
            joinedload(CommissionLog.order),
        )
    )
    if _dt_from:
        cl_q = cl_q.filter(CommissionLog.created_at >= _dt_from)
    if _dt_to:
        cl_q = cl_q.filter(CommissionLog.created_at <= _dt_to)
    commission_logs = cl_q.order_by(CommissionLog.created_at.desc()).all()

    cr_q = ClockRecord.query.options(joinedload(ClockRecord.user))
    if _dt_from:
        cr_q = cr_q.filter(ClockRecord.clock_in >= _dt_from)
    if _dt_to:
        cr_q = cr_q.filter(ClockRecord.clock_in <= _dt_to)
    clock_records = cr_q.order_by(ClockRecord.clock_in.desc()).all()

    intimacies = (
        Intimacy.query
        .options(
            joinedload(Intimacy.boss),
            joinedload(Intimacy.player),
        )
        .order_by(Intimacy.updated_at.desc())
        .all()
    )
    lotteries = (
        Lottery.query
        .options(joinedload(Lottery.creator))
        .order_by(Lottery.created_at.desc())
        .all()
    )
    lottery_winners = (
        LotteryWinner.query
        .options(
            joinedload(LotteryWinner.user),
            joinedload(LotteryWinner.lottery),
        )
        .order_by(LotteryWinner.created_at.desc())
        .all()
    )
    ol_q = OperationLog.query.options(joinedload(OperationLog.operator))
    if _dt_from:
        ol_q = ol_q.filter(OperationLog.created_at >= _dt_from)
    if _dt_to:
        ol_q = ol_q.filter(OperationLog.created_at <= _dt_to)
    operation_logs = ol_q.order_by(OperationLog.created_at.desc()).all()

    project_items = (
        ProjectItem.query
        .options(joinedload(ProjectItem.project))
        .order_by(ProjectItem.project_id.asc(), ProjectItem.sort_order.asc())
        .all()
    )

    user_rows = []
    for u in users:
        user_rows.append([
            u.id,
            _display_name(u),
            u.username,
            u.kook_username or '',
            u.kook_id or '',
            u.player_nickname or '',
            u.role_name,
            _identity_summary(u),
            '启用' if u.status else '禁用',
            _to_float(u.m_coin),
            _to_float(u.m_coin_gift),
            _to_float(u.m_bean),
            _to_float(u.m_bean_frozen),
            u.vip_level or '',
            _to_float(u.vip_discount),
            u.experience or 0,
            '是' if u.anonymous_broadcast_all else '否',
            u.register_type or '',
            fmt_dt(u.created_at, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(u.updated_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '用户详情',
        [
            '用户ID', '客户昵称', '用户名', 'KOOK名称', 'KOOK ID', '陪玩昵称',
            '主角色', '有效身份汇总', '状态',
            '嗯呢币余额', '赠金余额', '小猪粮余额', '冻结小猪粮',
            'VIP等级', '折扣(%)', '经验值', '全部匿名播报',
            '注册方式', '创建时间(北京)', '更新时间(北京)',
        ],
        user_rows,
    )
    sheets_summary.append((sheet, count, '用户基础资料 + 钱包 + 身份标签'))

    order_rows = []
    for o in orders:
        order_rows.append([
            o.id,
            o.order_no,
            o.status_label,
            o.order_type or '',
            o.boss_id,
            _display_name(o.boss),
            o.player_id,
            _display_name(o.player, prefer_player=True),
            o.staff_id or '',
            _display_name(o.staff),
            o.project_display,
            o.price_tier or '',
            _to_float(o.duration),
            _to_float(o.base_price),
            _to_float(o.extra_price),
            _to_float(o.addon_price),
            _to_float(o.total_price),
            _to_float(o.player_earning),
            _to_float(o.shop_earning),
            _to_float(o.commission_rate),
            _to_float(o.boss_discount),
            '是' if o.is_frozen else '否',
            fmt_dt(o.created_at, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(o.report_time, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(o.confirm_time, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(o.pay_time, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(o.refund_time, '%Y-%m-%d %H:%M:%S'),
            o.remark or '',
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '订单数据',
        [
            'ID', '订单号', '状态', '订单类型', '老板ID', '老板',
            '陪玩ID', '陪玩', '客服ID', '客服',
            '项目', '档位', '时长', '基础单价', '补充单价', '附加价格',
            '总价', '陪玩收益', '平台收益', '佣金比例(%)', '老板折扣(%)',
            '是否冻结', '创建时间(北京)', '申报时间(北京)', '确认时间(北京)',
            '支付时间(北京)', '退款时间(北京)', '备注',
        ],
        order_rows,
    )
    sheets_summary.append((sheet, count, '订单全流程明细（老板/陪玩/客服/金额/时间）'))

    gift_rows = []
    for g in gift_orders:
        gift_rows.append([
            g.id,
            g.status_label,
            g.boss_id,
            _display_name(g.boss),
            g.player_id,
            _display_name(g.player, prefer_player=True),
            g.staff_id or '',
            _display_name(g.staff),
            g.gift_id,
            g.gift.name if g.gift else '',
            g.quantity or 0,
            _to_float(g.unit_price),
            _to_float(g.total_price),
            _to_float(g.player_earning),
            _to_float(g.shop_earning),
            _to_float(g.commission_rate),
            '是' if g.is_frozen else '否',
            fmt_dt(g.created_at, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(g.refund_time, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '礼物订单',
        [
            'ID', '状态', '老板ID', '老板', '陪玩ID', '陪玩', '客服ID', '客服',
            '礼物ID', '礼物名', '数量', '单价', '总价', '陪玩收益', '平台收益',
            '佣金比例(%)', '是否冻结', '创建时间(北京)', '退款时间(北京)',
        ],
        gift_rows,
    )
    sheets_summary.append((sheet, count, '礼物赠送与退款明细'))

    staff_stats = {}
    player_stats = {}
    boss_stats = {}
    clock_stats = {}

    def _staff_bucket(uid):
        return staff_stats.setdefault(uid, {
            'dispatch_orders': 0,
            'paid_orders': 0,
            'pending_orders': 0,
            'refunded_orders': 0,
            'gift_dispatch_count': 0,
            'gift_dispatch_amount': 0.0,
        })

    def _player_bucket(uid):
        return player_stats.setdefault(uid, {
            'order_count': 0,
            'paid_order_count': 0,
            'refunded_order_count': 0,
            'paid_order_duration': 0.0,
            'paid_order_income': 0.0,
            'refunded_order_income': 0.0,
            'gift_count': 0,
            'gift_income': 0.0,
            'gift_refund_deduct': 0.0,
            'withdraw_total': 0.0,
        })

    def _boss_bucket(uid):
        return boss_stats.setdefault(uid, {
            'order_count': 0,
            'order_paid_amount': 0.0,
            'order_refund_amount': 0.0,
            'gift_count': 0,
            'gift_paid_amount': 0.0,
            'gift_refund_amount': 0.0,
            'recharge_total': 0.0,
            'consume_total': 0.0,
        })

    def _clock_bucket(uid):
        return clock_stats.setdefault(uid, {
            'days': set(),
            'minutes': 0,
            'last_clock_in': None,
        })

    for o in orders:
        if o.staff_id:
            sb = _staff_bucket(o.staff_id)
            sb['dispatch_orders'] += 1
            if o.status == 'paid':
                sb['paid_orders'] += 1
            elif o.status == 'refunded':
                sb['refunded_orders'] += 1
            elif o.status in ('pending_report', 'pending_confirm', 'pending_pay'):
                sb['pending_orders'] += 1

        pb = _player_bucket(o.player_id)
        pb['order_count'] += 1
        if o.status == 'paid':
            pb['paid_order_count'] += 1
            pb['paid_order_duration'] += _to_float(o.duration)
            pb['paid_order_income'] += _to_float(o.player_earning)
        elif o.status == 'refunded':
            pb['refunded_order_count'] += 1
            pb['refunded_order_income'] += _to_float(o.player_earning)

        bb = _boss_bucket(o.boss_id)
        bb['order_count'] += 1
        if o.status == 'paid':
            bb['order_paid_amount'] += _to_float(o.total_price)
        elif o.status == 'refunded':
            bb['order_refund_amount'] += _to_float(o.total_price)

    for g in gift_orders:
        if g.staff_id:
            sb = _staff_bucket(g.staff_id)
            sb['gift_dispatch_count'] += 1
            sb['gift_dispatch_amount'] += _to_float(g.total_price)

        pb = _player_bucket(g.player_id)
        pb['gift_count'] += 1
        if g.status == 'paid':
            pb['gift_income'] += _to_float(g.player_earning)
        elif g.status == 'refunded':
            pb['gift_refund_deduct'] += _to_float(g.player_earning)

        bb = _boss_bucket(g.boss_id)
        bb['gift_count'] += 1
        if g.status == 'paid':
            bb['gift_paid_amount'] += _to_float(g.total_price)
        elif g.status == 'refunded':
            bb['gift_refund_amount'] += _to_float(g.total_price)

    for wr in withdraw_requests:
        if wr.status in ('approved', 'paid'):
            pb = _player_bucket(wr.user_id)
            pb['withdraw_total'] += _to_float(wr.amount)

    for bl in balance_logs:
        bb = _boss_bucket(bl.user_id)
        amt = _to_float(bl.amount)
        if bl.change_type == 'recharge' and amt > 0:
            bb['recharge_total'] += amt
        if bl.change_type in ('consume', 'gift_send', 'admin_adjust') and amt < 0:
            bb['consume_total'] += abs(amt)

    for c in clock_records:
        cb = _clock_bucket(c.user_id)
        cb['minutes'] += c.duration_minutes or 0
        if c.clock_in:
            cb['days'].add(fmt_dt(c.clock_in, '%Y-%m-%d'))
            if not cb['last_clock_in'] or c.clock_in > cb['last_clock_in']:
                cb['last_clock_in'] = c.clock_in

    staff_rows = []
    for u in [x for x in users if x.is_staff]:
        sb = _staff_bucket(u.id)
        cb = _clock_bucket(u.id)
        staff_rows.append([
            u.id,
            _display_name(u),
            u.kook_username or '',
            _identity_summary(u),
            sb['dispatch_orders'],
            sb['paid_orders'],
            sb['pending_orders'],
            sb['refunded_orders'],
            sb['gift_dispatch_count'],
            round(sb['gift_dispatch_amount'], 2),
            len(cb['days']),
            round((cb['minutes'] or 0) / 60, 2),
            fmt_dt(cb['last_clock_in'], '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '客服数据',
        [
            '用户ID', '客服名称', 'KOOK名称', '有效身份汇总',
            '派单总数', '已结算订单', '待处理订单', '已退款订单',
            '礼物派发笔数', '礼物派发金额', '打卡天数', '累计工时(小时)',
            '最近上班时间(北京)',
        ],
        staff_rows,
    )
    sheets_summary.append((sheet, count, '客服/管理员维度运营数据'))

    player_rows = []
    for u in [x for x in users if x.has_player_tag]:
        pb = _player_bucket(u.id)
        player_rows.append([
            u.id,
            _display_name(u, prefer_player=True),
            u.kook_username or '',
            _identity_summary(u),
            pb['order_count'],
            pb['paid_order_count'],
            pb['refunded_order_count'],
            round(pb['paid_order_duration'], 2),
            round(pb['paid_order_income'], 2),
            round(pb['refunded_order_income'], 2),
            pb['gift_count'],
            round(pb['gift_income'], 2),
            round(pb['gift_refund_deduct'], 2),
            round(pb['withdraw_total'], 2),
            _to_float(u.m_bean),
            _to_float(u.m_coin),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '陪玩数据',
        [
            '用户ID', '陪玩名称', 'KOOK名称', '有效身份汇总',
            '接单总数', '已结算订单数', '已退款订单数', '已结算时长(小时)',
            '订单收益(小猪粮)', '退款扣减(订单)', '收礼笔数', '收礼收益(小猪粮)',
            '礼物退款扣减', '累计提现', '当前小猪粮', '当前嗯呢币',
        ],
        player_rows,
    )
    sheets_summary.append((sheet, count, '陪玩接单/收益/提现汇总'))

    boss_rows = []
    for u in [x for x in users if x.is_god]:
        bb = _boss_bucket(u.id)
        boss_rows.append([
            u.id,
            _display_name(u),
            u.kook_username or '',
            _identity_summary(u),
            bb['order_count'],
            round(bb['order_paid_amount'], 2),
            round(bb['order_refund_amount'], 2),
            bb['gift_count'],
            round(bb['gift_paid_amount'], 2),
            round(bb['gift_refund_amount'], 2),
            round(bb['recharge_total'], 2),
            round(bb['consume_total'], 2),
            _to_float(u.m_coin),
            _to_float(u.m_coin_gift),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '老板数据',
        [
            '用户ID', '老板名称', 'KOOK名称', '有效身份汇总',
            '订单总数', '订单已支付金额', '订单退款金额',
            '礼物总笔数', '礼物消费金额', '礼物退款金额',
            '累计充值', '累计消费', '当前嗯呢币', '当前赠金',
        ],
        boss_rows,
    )
    sheets_summary.append((sheet, count, '老板充值/消费/下单/赠礼汇总'))

    withdrawal_rows = []
    for wr in withdraw_requests:
        withdrawal_rows.append([
            wr.id,
            wr.user_id,
            _display_name(wr.user, prefer_player=True),
            _to_float(wr.amount),
            wr.status,
            wr.payment_method or '',
            wr.payment_account or '',
            wr.payment_image or '',
            wr.auditor_id or '',
            _display_name(wr.auditor),
            wr.audit_remark or '',
            fmt_dt(wr.created_at, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(wr.audit_at, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(wr.paid_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '提现数据',
        [
            'ID', '用户ID', '用户', '金额', '状态', '方式',
            '收款账号', '收款码图片', '审核人ID', '审核人',
            '审核备注', '申请时间(北京)', '审核时间(北京)', '打款时间(北京)',
        ],
        withdrawal_rows,
    )
    sheets_summary.append((sheet, count, '提现申请与审核流转'))

    balance_rows = []
    for bl in balance_logs:
        operator = user_map.get(bl.operator_id)
        balance_rows.append([
            bl.id,
            bl.user_id,
            _display_name(bl.user),
            bl.change_type,
            _to_float(bl.amount),
            _to_float(bl.balance_after),
            bl.reason or '',
            bl.operator_id or '',
            _display_name(operator),
            fmt_dt(bl.created_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '嗯呢币流水',
        [
            'ID', '用户ID', '用户', '变动类型', '变动金额',
            '变动后余额', '原因', '操作人ID', '操作人', '时间(北京)',
        ],
        balance_rows,
    )
    sheets_summary.append((sheet, count, '嗯呢币余额变动明细'))

    commission_rows = []
    for cl in commission_logs:
        commission_rows.append([
            cl.id,
            cl.user_id,
            _display_name(cl.user, prefer_player=True),
            cl.change_type,
            _to_float(cl.amount),
            _to_float(cl.balance_after),
            cl.order_id or '',
            cl.order.order_no if cl.order else '',
            cl.reason or '',
            fmt_dt(cl.created_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '小猪粮流水',
        [
            'ID', '用户ID', '用户', '变动类型', '变动金额', '变动后余额',
            '关联订单ID', '关联订单号', '原因', '时间(北京)',
        ],
        commission_rows,
    )
    sheets_summary.append((sheet, count, '小猪粮收益/提现/扣减明细'))

    clock_rows = []
    for c in clock_records:
        clock_rows.append([
            c.id,
            c.user_id,
            _display_name(c.user),
            _identity_summary(c.user) if c.user else '',
            fmt_dt(c.clock_in, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(c.clock_out, '%Y-%m-%d %H:%M:%S'),
            round((c.duration_minutes or 0) / 60, 2),
            c.status_label,
            c.remark or '',
            fmt_dt(c.created_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '打卡数据',
        [
            'ID', '用户ID', '用户', '身份', '上班时间(北京)', '下班时间(北京)',
            '工时(小时)', '状态', '备注', '记录创建时间(北京)',
        ],
        clock_rows,
    )
    sheets_summary.append((sheet, count, '客服/管理员打卡明细'))

    intimacy_rows = []
    for rel in intimacies:
        intimacy_rows.append([
            rel.id,
            rel.boss_id,
            _display_name(rel.boss),
            rel.player_id,
            _display_name(rel.player, prefer_player=True),
            _to_float(rel.value),
            fmt_dt(rel.created_at, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(rel.updated_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '亲密度关系',
        [
            'ID', '老板ID', '老板', '陪玩ID', '陪玩',
            '亲密度', '创建时间(北京)', '更新时间(北京)',
        ],
        intimacy_rows,
    )
    sheets_summary.append((sheet, count, '老板-陪玩亲密度关系'))

    lottery_rows = []
    for l in lotteries:
        lottery_rows.append([
            l.id,
            l.mode_label,
            l.title,
            l.description or '',
            l.prize,
            l.winner_count,
            l.participants.count() if l.is_interactive else '',
            l.status_label,
            l.channel_id,
            l.kook_msg_id or '',
            l.emoji or '',
            '、'.join(l.get_eligible_roles()),
            l.min_vip_level or '',
            l.created_by,
            _display_name(l.creator),
            fmt_dt(l.draw_time, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(l.created_at, '%Y-%m-%d %H:%M:%S'),
            fmt_dt(l.updated_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '抽奖活动',
        [
            'ID', '类型', '标题', '简介', '奖品', '中奖人数', '参与人数', '状态', '频道ID',
            '消息ID', '表情', '可参与角色', '最低VIP', '创建人ID', '创建人',
            '开奖时间(北京)', '创建时间(北京)', '更新时间(北京)',
        ],
        lottery_rows,
    )
    sheets_summary.append((sheet, count, '抽奖配置与活动状态'))

    winner_rows = []
    for w in lottery_winners:
        winner_rows.append([
            w.id,
            w.lottery_id,
            w.lottery.title if w.lottery else '',
            w.user_id or '',
            _display_name(w.user),
            w.kook_id,
            '是' if w.is_rigged else '否',
            '是' if w.notified else '否',
            fmt_dt(w.created_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '抽奖中奖记录',
        [
            'ID', '抽奖ID', '抽奖标题', '用户ID', '用户', 'KOOK ID',
            '是否内定', '是否已通知', '创建时间(北京)',
        ],
        winner_rows,
    )
    sheets_summary.append((sheet, count, '抽奖中奖名单明细'))

    op_rows = []
    for log in operation_logs:
        op_rows.append([
            log.id,
            log.operator_id,
            log.operator_display_name,
            log.action_type,
            log.target_type or '',
            log.target_id or '',
            log.detail or '',
            fmt_dt(log.created_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '操作日志',
        [
            'ID', '操作人ID', '操作人', '操作类型', '目标类型', '目标ID',
            '详情', '时间(北京)',
        ],
        op_rows,
    )
    sheets_summary.append((sheet, count, '后台操作审计日志'))

    project_rows = []
    for item in project_items:
        project_rows.append([
            item.id,
            item.project_id,
            item.project.name if item.project else '',
            item.name,
            _to_float(item.price_casual),
            _to_float(item.price_tech),
            _to_float(item.price_god),
            _to_float(item.price_pro),
            _to_float(item.price_devil),
            _to_float(item.commission_rate),
            item.billing_type or '',
            item.project_type or '',
            item.sort_order or 0,
            '启用' if item.status else '停用',
            fmt_dt(item.created_at, '%Y-%m-%d %H:%M:%S'),
        ])
    sheet, count = _append_sheet(
        wb,
        used_sheet_names,
        '项目配置',
        [
            'ID', '游戏ID', '游戏名', '子项目',
            '娱乐价', '技术价', '大神价', '巅峰价', '魔王价',
            '佣金比例(%)', '计费类型', '项目类型', '排序', '状态', '创建时间(北京)',
        ],
        project_rows,
    )
    sheets_summary.append((sheet, count, '游戏项目与档位价格'))

    _style_header(info_ws, ['导出时间(北京)', 'Sheet', '记录数', '说明'])
    export_time = fmt_dt(datetime.utcnow(), '%Y-%m-%d %H:%M:%S')
    for idx, (sheet_name, row_count, desc) in enumerate(sheets_summary, start=2):
        info_ws.cell(row=idx, column=1, value=export_time)
        info_ws.cell(row=idx, column=2, value=sheet_name)
        info_ws.cell(row=idx, column=3, value=row_count)
        info_ws.cell(row=idx, column=4, value=desc)

    if include_sections:
        selected_keys = {str(x).strip() for x in include_sections if str(x).strip() in EXPORT_SECTION_KEYS}
        allowed_sheets = {'导出说明'}
        for item in EXPORT_SECTION_OPTIONS:
            if item['key'] in selected_keys:
                allowed_sheets.add(item['sheet'])

        for sheet_name in list(wb.sheetnames):
            if sheet_name not in allowed_sheets:
                wb.remove(wb[sheet_name])

        # 重建导出说明（仅保留已勾选内容）
        info_ws = wb['导出说明']
        if info_ws.max_row > 0:
            info_ws.delete_rows(1, info_ws.max_row)
        _style_header(info_ws, ['导出时间(北京)', 'Sheet', '记录数', '说明'])
        export_time = fmt_dt(datetime.utcnow(), '%Y-%m-%d %H:%M:%S')
        row_idx = 2
        for item in EXPORT_SECTION_OPTIONS:
            if item['key'] not in selected_keys or item['sheet'] not in wb.sheetnames:
                continue
            ws = wb[item['sheet']]
            info_ws.cell(row=row_idx, column=1, value=export_time)
            info_ws.cell(row=row_idx, column=2, value=item['sheet'])
            info_ws.cell(row=row_idx, column=3, value=max(ws.max_row - 1, 0))
            info_ws.cell(row=row_idx, column=4, value=item['desc'])
            row_idx += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
