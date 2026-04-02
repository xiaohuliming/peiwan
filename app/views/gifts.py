from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import func, or_, case
from decimal import Decimal

from app.models.gift import Gift, GiftOrder
from app.models.user import User
from app.extensions import db
from app.services import gift_service
from app.services import kook_service
from app.utils.permissions import staff_required

gifts_bp = Blueprint('gifts', __name__)


@gifts_bp.route('/')
@login_required
def index():
    """礼物记录列表"""
    if not current_user.is_staff:
        flash('无权访问', 'error')
        return redirect(url_for('dashboard.index'))

    page = request.args.get('page', 1, type=int)
    query = GiftOrder.query

    # 筛选: 礼物名称
    gift_name = request.args.get('gift_name', '').strip()
    if gift_name:
        gift_ids = db.session.query(Gift.id).filter(Gift.name.ilike(f'%{gift_name}%')).subquery()
        query = query.filter(GiftOrder.gift_id.in_(gift_ids))

    # 筛选: 老板昵称
    boss_name = request.args.get('boss_name', '').strip()
    if boss_name:
        boss_ids = db.session.query(User.id).filter(
            User.role_filter_expr('god'), User.nickname.ilike(f'%{boss_name}%')
        ).subquery()
        query = query.filter(GiftOrder.boss_id.in_(boss_ids))

    # 筛选: 收礼人昵称（陪玩/老板）
    player_name = request.args.get('player_name', '').strip()
    if player_name:
        player_ids = db.session.query(User.id).filter(
            or_(
                User.role_filter_expr('player'),
                User.role_filter_expr('god')
            ),
            or_(
                User.player_nickname.ilike(f'%{player_name}%'),
                User.nickname.ilike(f'%{player_name}%')
            )
        ).subquery()
        query = query.filter(GiftOrder.player_id.in_(player_ids))

    # 筛选: 派单客服
    staff_name = request.args.get('staff_name', '').strip()
    if staff_name:
        staff_ids = db.session.query(User.id).filter(
            or_(
                User.player_nickname.ilike(f'%{staff_name}%'),
                User.nickname.ilike(f'%{staff_name}%')
            )
        ).subquery()
        query = query.filter(GiftOrder.staff_id.in_(staff_ids))

    # 日期范围
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    if date_from:
        query = query.filter(GiftOrder.created_at >= date_from)
    if date_to:
        query = query.filter(GiftOrder.created_at <= date_to + ' 23:59:59')

    # 统计卡片
    total_amount = db.session.query(func.sum(GiftOrder.total_price)).filter(
        GiftOrder.status == 'paid'
    ).scalar() or Decimal('0')
    player_wages = db.session.query(func.sum(GiftOrder.player_earning)).filter(
        GiftOrder.status == 'paid'
    ).scalar() or Decimal('0')
    platform_revenue = db.session.query(func.sum(GiftOrder.shop_earning)).filter(
        GiftOrder.status == 'paid'
    ).scalar() or Decimal('0')

    stats = {
        'total_amount': total_amount,
        'player_wages': player_wages,
        'platform_revenue': platform_revenue,
    }

    frozen_first = case((GiftOrder.freeze_status == 'frozen', 1), else_=0)
    gift_orders = query.order_by(
        frozen_first.desc(),
        GiftOrder.created_at.desc(),
    ).paginate(page=page, per_page=15)
    pagination_args = request.args.to_dict(flat=True)
    pagination_args.pop('page', None)

    return render_template(
        'gifts/index.html',
        gift_orders=gift_orders,
        stats=stats,
        current_subtab='gift',
        pagination_args=pagination_args,
    )


@gifts_bp.route('/send', methods=['GET', 'POST'])
@login_required
@staff_required
def send():
    """赠送礼物"""
    if request.method == 'POST':
        gift_id = request.form.get('gift_id', type=int)
        boss_id = request.form.get('boss_id', type=int)
        player_id = request.form.get('player_id', type=int)
        quantity = request.form.get('quantity', 1, type=int)

        if not gift_id or not boss_id or not player_id:
            flash('请填写完整信息', 'error')
            return redirect(url_for('gifts.send'))

        gift = Gift.query.get(gift_id)
        boss = User.query.get(boss_id)
        player = User.query.get(player_id)

        if not gift or not boss or not player:
            flash('数据不存在', 'error')
            return redirect(url_for('gifts.send'))

        if not gift.status:
            flash('该礼物已下架', 'error')
            return redirect(url_for('gifts.send'))

        if not boss.is_god:
            flash('老板角色错误', 'error')
            return redirect(url_for('gifts.send'))

        # 收礼人支持：陪玩身份 或 老板身份（含身份标签）
        if not (player.is_player or player.is_god):
            flash('收礼人身份错误（需有陪玩或老板身份）', 'error')
            return redirect(url_for('gifts.send'))

        gift_order, error = gift_service.send_gift(
            boss=boss, player=player, gift=gift,
            quantity=quantity, staff=current_user
        )

        if error:
            flash(error, 'error')
            return redirect(url_for('gifts.send'))

        db.session.commit()
        # KOOK 推送: 私信陪玩 + 礼物频道播报
        kook_service.push_gift_to_player(gift_order)
        kook_service.push_gift_broadcast(gift_order)
        flash(f'礼物赠送成功: {gift.name} x{quantity}', 'success')
        return redirect(url_for('gifts.index'))

    gifts = Gift.query.filter_by(status=True).order_by(Gift.sort_order.asc(), Gift.id.asc()).all()
    return render_template('gifts/send.html', gifts=gifts)


@gifts_bp.route('/<int:order_id>/action/<action>', methods=['POST'])
@login_required
@staff_required
def gift_action(order_id, action):
    """冻结/解冻/退款"""
    gift_order = GiftOrder.query.get_or_404(order_id)

    if action == 'freeze':
        success, error = gift_service.freeze_gift_order(gift_order, current_user.id)
        if success:
            db.session.commit()
            flash('礼物订单已冻结', 'success')
        else:
            flash(error, 'error')

    elif action == 'unfreeze':
        success, error = gift_service.unfreeze_gift_order(gift_order, current_user.id)
        if success:
            db.session.commit()
            flash('礼物订单已解冻', 'success')
        else:
            flash(error, 'error')

    elif action == 'refund':
        if not current_user.is_staff:
            flash('退款需要客服及以上权限', 'error')
            return redirect(url_for('gifts.index'))
        success, error = gift_service.refund_gift_order(gift_order, current_user.id)
        if success:
            db.session.commit()
            # KOOK 推送: 礼物退款后私信老板和陪玩
            kook_service.push_gift_refund_notice(gift_order)
            flash('礼物退款成功', 'success')
        else:
            flash(error, 'error')

    else:
        flash('未知操作', 'error')

    return redirect(request.referrer or url_for('gifts.index'))


@gifts_bp.route('/export')
@login_required
@staff_required
def export_gifts():
    """导出筛选后的礼物记录为 Excel"""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        flash('导出失败，请安装 openpyxl', 'error')
        return redirect(url_for('gifts.index'))

    from app.utils.time_utils import to_beijing
    from datetime import datetime as dt_cls

    query = GiftOrder.query

    gift_name = request.args.get('gift_name', '').strip()
    if gift_name:
        gift_ids = db.session.query(Gift.id).filter(Gift.name.ilike(f'%{gift_name}%')).subquery()
        query = query.filter(GiftOrder.gift_id.in_(gift_ids))

    boss_name = request.args.get('boss_name', '').strip()
    if boss_name:
        boss_ids = db.session.query(User.id).filter(
            User.role_filter_expr('god'), User.nickname.ilike(f'%{boss_name}%')
        ).subquery()
        query = query.filter(GiftOrder.boss_id.in_(boss_ids))

    player_name = request.args.get('player_name', '').strip()
    if player_name:
        player_ids = db.session.query(User.id).filter(
            or_(
                User.role_filter_expr('player'),
                User.role_filter_expr('god')
            ),
            or_(
                User.player_nickname.ilike(f'%{player_name}%'),
                User.nickname.ilike(f'%{player_name}%')
            )
        ).subquery()
        query = query.filter(GiftOrder.player_id.in_(player_ids))

    staff_name = request.args.get('staff_name', '').strip()
    if staff_name:
        staff_ids = db.session.query(User.id).filter(
            or_(
                User.player_nickname.ilike(f'%{staff_name}%'),
                User.nickname.ilike(f'%{staff_name}%')
            )
        ).subquery()
        query = query.filter(GiftOrder.staff_id.in_(staff_ids))

    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    if date_from:
        from datetime import datetime as dt_mod, timezone as tz_mod, timedelta as td_mod
        _bj = tz_mod(td_mod(hours=8))
        try:
            utc_from = dt_mod.strptime(date_from, '%Y-%m-%d').replace(tzinfo=_bj).astimezone(tz_mod.utc).replace(tzinfo=None)
            query = query.filter(GiftOrder.created_at >= utc_from)
        except ValueError:
            pass
    if date_to:
        from datetime import datetime as dt_mod, timezone as tz_mod, timedelta as td_mod
        _bj = tz_mod(td_mod(hours=8))
        try:
            utc_to = dt_mod.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S').replace(tzinfo=_bj).astimezone(tz_mod.utc).replace(tzinfo=None)
            query = query.filter(GiftOrder.created_at <= utc_to)
        except ValueError:
            pass

    rows = query.order_by(GiftOrder.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '礼物记录'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    headers = ['老板', '收礼人', '礼物', '类型', '数量', '总价', '收礼收益',
               '平台营收', '佣金比例', '客服', '状态', '时间']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    status_map = {'paid': '已完成', 'refunded': '已退款'}
    type_map = {'normal': '普通', 'crown': '冠名'}

    def _fmt(dt_val):
        bj = to_beijing(dt_val)
        return bj.strftime('%Y-%m-%d %H:%M:%S') if bj else ''

    for i, g in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=(g.boss.nickname or g.boss.username) if g.boss else '-')
        ws.cell(row=i, column=2, value=(g.player.player_nickname or g.player.nickname or g.player.username) if g.player else '-')
        ws.cell(row=i, column=3, value=g.gift.name if g.gift else '-')
        ws.cell(row=i, column=4, value=type_map.get(g.gift.gift_type, g.gift.gift_type) if g.gift else '-')
        ws.cell(row=i, column=5, value=g.quantity or 0)
        ws.cell(row=i, column=6, value=float(g.total_price or 0))
        ws.cell(row=i, column=7, value=float(g.player_earning or 0))
        ws.cell(row=i, column=8, value=float(g.shop_earning or 0))
        ws.cell(row=i, column=9, value=f'{g.commission_rate}%' if g.commission_rate else '-')
        ws.cell(row=i, column=10, value=g.staff.staff_display_name if g.staff else '-')
        frozen_label = '冻结中' if g.is_frozen else ''
        ws.cell(row=i, column=11, value=frozen_label or status_map.get(g.status, g.status))
        ws.cell(row=i, column=12, value=_fmt(g.created_at))

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value or ''))
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    filename = f'礼物记录_{dt_cls.now().strftime("%Y%m%d")}.xlsx'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
